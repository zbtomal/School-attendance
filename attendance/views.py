from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from concurrent.futures import ThreadPoolExecutor
from django.conf import settings
from .models import Teacher, Student, Attendance, TeacherAttendance
from django.urls import reverse
from .sms_utils import build_absent_message, build_teacher_absent_message, send_sms
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Expected Excel column headers (in order) for the two bulk-upload forms.
# These must exactly match (case/whitespace-insensitive) the header row
# shown to the admin in student_upload.html / teacher_upload.html.
STUDENT_EXCEL_HEADERS = ["Roll", "Name", "Class", "Section", "Session", "Phone"]
TEACHER_EXCEL_HEADERS = ["Name", "Number", "Class"]


def send_absent_sms(people, message_builder, date_str, number_getter):
    """Send absence alerts concurrently so one slow SMS does not block every other alert."""
    people_with_numbers = [person for person in people if number_getter(person)]
    people_without_numbers = [f"{person.name} (no phone)" for person in people if not number_getter(person)]

    def deliver(person):
        message = message_builder(person.name, date_str)
        success, response = send_sms(number_getter(person), message)
        return person.name if success else None

    if not people_with_numbers:
        return 0, people_without_numbers

    worker_count = min(8, len(people_with_numbers))
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        results = list(executor.map(deliver, people_with_numbers))

    sent_count = sum(1 for result in results if result)
    failed = people_without_numbers + [
        person.name for person, result in zip(people_with_numbers, results) if not result
    ]
    return sent_count, failed


def is_admin(user):
    return user.is_staff


def check_header(header_row, expected_headers):
    """
    Compares the first row of an uploaded Excel file against the expected
    column headers (case/whitespace-insensitive, order matters).
    Returns None if it matches, otherwise a human-readable error message.
    """
    cells = list(header_row) if header_row else []
    actual = []
    for i in range(len(expected_headers)):
        cell = cells[i] if i < len(cells) else None
        actual.append(str(cell).strip().lower() if cell is not None else "")
    expected = [h.strip().lower() for h in expected_headers]

    if actual != expected:
        expected_display = ", ".join(expected_headers)
        return (
            f"Header row is not in the correct format. The first row must have "
            f"these exact columns, in this order: {expected_display}. "
            f"Please fix the header and upload again."
        )
    return None


def get_class_choices():
    import re

    names = set()
    for c in Student.objects.values_list('class_name', flat=True):
        if c is None:
            continue
        cleaned = str(c).strip()
        if cleaned:
            names.add(cleaned)

    def sort_key(name):
        s = str(name).strip()
        low = s.lower()

        if low.startswith(('play', 'nursery', 'kg', 'pre')):
            group = 0
        elif any(ch.isdigit() for ch in s):
            group = 1
        else:
            group = 2

        m = re.search(r'(\d+)', s)
        num = int(m.group(1)) if m else 0
        return (group, num, low)

    return sorted(names, key=sort_key)


def get_section_choices():
    return list(
        Student.objects.exclude(section='').values_list('section', flat=True).distinct().order_by('section')
    )


def build_choice_options(values, selected_value):
    return [{'value': v, 'is_selected': (str(v) == str(selected_value))} for v in values]


def build_choice_options_multi(values, selected_values):
    selected_set = {str(v) for v in selected_values}
    return [{'value': v, 'is_selected': (str(v) in selected_set)} for v in values]


@login_required
def dashboard(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    total_students = 0
    present_count = 0
    absent_count = 0
    today = timezone.now().date()
    teacher_classes = []

    if teacher:
        teacher_classes = teacher.get_class_list()
        students = Student.objects.filter(class_name__in=teacher_classes)
        total_students = students.count()
        attendance_today = Attendance.objects.filter(
            student__class_name__in=teacher_classes,
            date=today
        )
        present_count = attendance_today.filter(is_present=True).count()
        absent_count = attendance_today.filter(is_present=False).count()

    if request.user.is_staff:
        all_today = Attendance.objects.filter(date=today)
        admin_present = all_today.filter(is_present=True).count()
        admin_absent = all_today.filter(is_present=False).count()
        admin_total = Student.objects.count()
    else:
        admin_present = admin_absent = admin_total = 0

    return render(request, 'attendance/dashboard.html', {
        'teacher': teacher,
        'teacher_classes': teacher_classes,
        'total_students': total_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'today': today,
        'admin_present': admin_present,
        'admin_absent': admin_absent,
        'admin_total': admin_total,
    })


@login_required
def change_password(request):
    error = None
    success = False

    if request.method == 'POST':
        old_password = request.POST.get('old_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if not request.user.check_password(old_password):
            error = "Current password is incorrect."
        elif len(new_password) < 4:
            error = "New password must be at least 4 characters."
        elif new_password != confirm_password:
            error = "New passwords do not match."
        else:
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            success = True

    return render(request, 'attendance/change_password.html', {'error': error, 'success': success})


@login_required
def mark_attendance(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    is_admin_user = request.user.is_staff
    selected_class = request.GET.get('class', '').strip()

    if not is_admin_user and not teacher:
        return render(request, 'attendance/mark_attendance.html', {
            'error': 'Your account is not linked to any teacher. Ask admin to link it.',
            'is_admin_user': False,
        })

    if is_admin_user:
        class_choices = get_class_choices()
    else:
        class_choices = teacher.get_class_list()
        if not class_choices:
            return render(request, 'attendance/mark_attendance.html', {
                'error': 'No class has been assigned to you yet. Please ask the admin to assign one.',
                'is_admin_user': False,
            })

    show_class_selector = is_admin_user or len(class_choices) > 1

    if not is_admin_user and len(class_choices) == 1:
        current_class = class_choices[0]
    elif selected_class in class_choices:
        current_class = selected_class
    else:
        current_class = None

    all_classes = [{'name': c, 'is_selected': (c == current_class)} for c in class_choices]

    def load_students(class_name):
        if is_admin_user:
            return Student.objects.filter(class_name=class_name).order_by('section', 'roll_no')
        return Student.objects.filter(class_name=class_name).order_by('roll_no')

    students = load_students(current_class) if current_class else Student.objects.none()

    today = timezone.now().date()
    date_str = today.strftime("%d-%b-%y")

    already_marked = False
    if current_class:
        already_marked = Attendance.objects.filter(
            student__class_name=current_class,
            date=today
        ).exists()

    saved = False
    sms_sent_count = 0
    sms_warning = None

    if request.method == 'POST' and current_class and not already_marked:
        post_class = request.POST.get('class', '').strip()
        if post_class and post_class in class_choices and post_class != current_class:
            current_class = post_class
            students = load_students(current_class)
            all_classes = [{'name': c, 'is_selected': (c == current_class)} for c in class_choices]
            already_marked = Attendance.objects.filter(
                student__class_name=current_class, date=today
            ).exists()

        if not already_marked:
            absent_students = []
            for student in students:
                status = request.POST.get(f'att_{student.id}', 'present')
                is_present = status == 'present'

                Attendance.objects.update_or_create(
                    student=student,
                    date=today,
                    defaults={'is_present': is_present}
                )

                if not is_present:
                    absent_students.append(student)

            sms_sent_count, sms_failed = send_absent_sms(
                absent_students,
                build_absent_message,
                date_str,
                lambda student: student.parent_mobile,
            )

            saved = True
            already_marked = True
            if sms_failed:
                sms_warning = f"SMS could not be sent to: {', '.join(sms_failed)}"

    attendance_map = {}
    if current_class:
        attendance_map = {
            a.student_id: a.is_present
            for a in Attendance.objects.filter(student__class_name=current_class, date=today)
        }

    student_rows = [
        {
            'student': s,
            'is_present': attendance_map.get(s.id, True),
        }
        for s in students
    ]

    if already_marked and current_class:
        present_count = Attendance.objects.filter(
            student__class_name=current_class, date=today, is_present=True
        ).count()
        absent_count = Attendance.objects.filter(
            student__class_name=current_class, date=today, is_present=False
        ).count()
    else:
        present_count = 0
        absent_count = 0

    return render(request, 'attendance/mark_attendance.html', {
        'teacher': teacher,
        'students': students,
        'student_rows': student_rows,
        'today': today,
        'saved': saved,
        'sms_sent_count': sms_sent_count,
        'sms_warning': sms_warning,
        'is_admin_user': is_admin_user,
        'show_class_selector': show_class_selector,
        'all_classes': all_classes,
        'current_class': current_class,
        'already_marked': already_marked,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_students': students.count() if hasattr(students, 'count') else len(students),
    })


@login_required
@user_passes_test(is_admin)
def student_list(request):
    query = request.GET.get('q', '').strip()
    class_filter = request.GET.get('class', '').strip()

    class_names = get_class_choices()
    all_classes = [{'name': c, 'is_selected': (str(c) == class_filter)} for c in class_names]

    students = Student.objects.all().order_by('class_name', 'section', 'roll_no')
    if class_filter:
        students = students.filter(class_name=class_filter)
    if query:
        students = students.filter(name__icontains=query)

    total_count = students.count()

    paginator = Paginator(students, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'attendance/student_list.html', {
        'students': page_obj,
        'total_count': total_count,
        'query': query,
        'class_filter': class_filter,
        'all_classes': all_classes,
    })


@login_required
@user_passes_test(is_admin)
def student_add(request):
    error = None
    if request.method == 'POST':
        try:
            Student.objects.create(
                roll_no=request.POST.get('roll_no', '').strip(),
                name=request.POST.get('name', '').strip(),
                class_name=request.POST.get('class_name', '').strip(),
                section=request.POST.get('section', '').strip(),
                parent_mobile=request.POST.get('parent_mobile', '').strip(),
            )
            return redirect('student_list')
        except IntegrityError:
            error = "A student with this Roll Number already exists in this Class/Section. Please check and try again."

    return render(request, 'attendance/student_form.html', {
        'mode': 'Add',
        'error': error,
        'class_options': build_choice_options(get_class_choices(), ''),
        'section_options': build_choice_options(get_section_choices(), ''),
    })


@login_required
@user_passes_test(is_admin)
def student_edit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    error = None
    if request.method == 'POST':
        try:
            student.roll_no = request.POST.get('roll_no', '').strip()
            student.name = request.POST.get('name', '').strip()
            student.class_name = request.POST.get('class_name', '').strip()
            student.section = request.POST.get('section', '').strip()
            student.parent_mobile = request.POST.get('parent_mobile', '').strip()
            student.save()
            return redirect('student_list')
        except IntegrityError:
            error = "A student with this Roll Number already exists in this Class/Section. Please check and try again."

    return render(request, 'attendance/student_form.html', {
        'mode': 'Edit',
        'student': student,
        'error': error,
        'class_options': build_choice_options(get_class_choices(), student.class_name),
        'section_options': build_choice_options(get_section_choices(), student.section),
    })


@login_required
@user_passes_test(is_admin)
def student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'attendance/student_confirm_delete.html', {'student': student})


@login_required
@user_passes_test(is_admin)
def class_delete(request, class_name):
    students = Student.objects.filter(class_name=class_name)
    count = students.count()

    if request.method == 'POST':
        students.delete()
        return redirect('student_list')

    return render(request, 'attendance/class_confirm_delete.html', {
        'class_name': class_name,
        'count': count,
    })


@login_required
@user_passes_test(is_admin)
def student_upload(request):
    file_results = []

    if request.method == 'POST' and request.FILES.getlist('excel_file'):
        for excel_file in request.FILES.getlist('excel_file'):

            if not excel_file.name.lower().endswith('.xlsx'):
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This is not an .xlsx file. Please save it as an Excel (.xlsx) file and try again.",
                })
                continue

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Could not open this file. It may be corrupted, password-protected, or not a real Excel file.",
                })
                continue

            if not rows:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, or no rows at all).",
                })
                continue

            header_error = check_header(rows[0], STUDENT_EXCEL_HEADERS)
            if header_error:
                file_results.append({
                    "filename": excel_file.name,
                    "error": header_error,
                })
                continue

            if len(rows) < 2:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, or no rows at all).",
                })
                continue

            parsed_rows = {}
            skipped = 0
            detected_classes = set()

            for row in rows[1:]:
                # Ignore completely blank/empty rows silently
                if not row or not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue

                if len(row) < 3:
                    skipped += 1
                    continue

                roll_raw = row[0]
                name_raw = row[1]
                class_raw = row[2]
                section_raw = row[3] if len(row) > 3 and row[3] is not None else ""
                phone_raw = row[5] if len(row) > 5 and row[5] is not None else ""

                roll_str = str(roll_raw).strip() if roll_raw is not None else ""
                if roll_str.endswith(".0"):
                    roll_str = roll_str[:-2]

                class_str = str(class_raw).strip() if class_raw is not None else ""
                if class_str.endswith(".0"):
                    class_str = class_str[:-2]

                name_str = str(name_raw).strip() if name_raw is not None else ""
                section_str = str(section_raw).strip()
                if section_str.endswith(".0"):
                    section_str = section_str[:-2]

                phone_str = str(phone_raw).strip()
                if phone_str.endswith(".0"):
                    phone_str = phone_str[:-2]
                if phone_str and not phone_str.startswith("0") and phone_str.isdigit():
                    phone_str = "0" + phone_str

                if not roll_str or not name_str or not class_str:
                    skipped += 1
                    continue

                key = (class_str, section_str, roll_str)
                parsed_rows[key] = {
                    "roll_no": roll_str,
                    "class_name": class_str,
                    "section": section_str,
                    "name": name_str,
                    "parent_mobile": phone_str,
                }
                detected_classes.add(f"{class_str}{section_str}")

            if not parsed_rows:
                file_results.append({
                    "filename": excel_file.name,
                    "error": f"No valid rows found. All {skipped} row(s) were missing a Roll, Name, or Class value.",
                })
                continue

            # Batch lookup existing students for detected classes in 1 query
            classes_in_file = {k[0] for k in parsed_rows.keys()}
            existing_students = {
                (s.class_name, s.section, s.roll_no): s
                for s in Student.objects.filter(class_name__in=classes_in_file)
            }

            to_create = []
            to_update = []
            created = 0
            updated = 0

            for key, data in parsed_rows.items():
                if key in existing_students:
                    s = existing_students[key]
                    changed = False
                    if s.name != data["name"]:
                        s.name = data["name"]
                        changed = True
                    if s.parent_mobile != data["parent_mobile"]:
                        s.parent_mobile = data["parent_mobile"]
                        changed = True
                    if changed:
                        to_update.append(s)
                    updated += 1
                else:
                    new_s = Student(
                        class_name=data["class_name"],
                        section=data["section"],
                        roll_no=data["roll_no"],
                        name=data["name"],
                        parent_mobile=data["parent_mobile"],
                    )
                    to_create.append(new_s)
                    existing_students[key] = new_s
                    created += 1

            with transaction.atomic():
                if to_create:
                    Student.objects.bulk_create(to_create, batch_size=500)
                if to_update:
                    Student.objects.bulk_update(to_update, fields=["name", "parent_mobile"], batch_size=500)

            file_results.append({
                "filename": excel_file.name,
                "class_label": ", ".join(sorted(detected_classes)) or "Unknown",
                "created": created,
                "updated": updated,
                "skipped": skipped,
            })

    return render(request, 'attendance/student_upload.html', {'file_results': file_results})


@login_required
@user_passes_test(is_admin)
def teacher_list(request):
    full_time = Teacher.objects.select_related('user').filter(
        employment_type=Teacher.EMPLOYMENT_FULL
    ).order_by('id')
    part_time = Teacher.objects.select_related('user').filter(
        employment_type=Teacher.EMPLOYMENT_PART
    ).order_by('id')
    return render(request, 'attendance/teacher_list.html', {
        'full_time_teachers': full_time,
        'part_time_teachers': part_time,
        'full_time_count': full_time.count(),
    })


@login_required
@user_passes_test(is_admin)
def teacher_add(request):
    error = None
    employment_type = request.GET.get('type', '').strip() or request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
    if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
        employment_type = Teacher.EMPLOYMENT_FULL

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        name = request.POST.get('name', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        assigned_classes_list = request.POST.getlist('assigned_classes')
        assigned_classes = ",".join(c.strip() for c in assigned_classes_list if c.strip())

        if User.objects.filter(username=username).exists():
            error = f"Username '{username}' is already taken."
        else:
            user = User.objects.create_user(username=username, password=password)
            Teacher.objects.create(
                user=user,
                name=name,
                mobile=mobile,
                assigned_classes=assigned_classes,
                employment_type=employment_type,
            )
            return redirect('teacher_list')

    type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'
    return render(request, 'attendance/teacher_form.html', {
        'mode': 'Add',
        'error': error,
        'class_options': build_choice_options_multi(get_class_choices(), []),
        'employment_type': employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        teacher.name = request.POST.get('name', '').strip()
        teacher.mobile = request.POST.get('mobile', '').strip()
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        teacher.employment_type = employment_type
        assigned_classes_list = request.POST.getlist('assigned_classes')
        teacher.assigned_classes = ",".join(c.strip() for c in assigned_classes_list if c.strip())
        teacher.save()

        new_password = request.POST.get('password', '').strip()
        if new_password and teacher.user:
            teacher.user.set_password(new_password)
            teacher.user.save()

        return redirect('teacher_list')

    type_label = 'Full-time' if teacher.employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'
    return render(request, 'attendance/teacher_form.html', {
        'mode': 'Edit',
        'teacher': teacher,
        'class_options': build_choice_options_multi(get_class_choices(), teacher.get_class_list()),
        'employment_type': teacher.employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_upload(request):
    file_results = []
    employment_type = request.GET.get('type', request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL)).strip()
    if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
        employment_type = Teacher.EMPLOYMENT_FULL
    type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'

    if request.method == 'POST' and request.FILES.getlist('excel_file'):
        employment_type = request.POST.get('employment_type', Teacher.EMPLOYMENT_FULL).strip()
        if employment_type not in (Teacher.EMPLOYMENT_FULL, Teacher.EMPLOYMENT_PART):
            employment_type = Teacher.EMPLOYMENT_FULL
        type_label = 'Full-time' if employment_type == Teacher.EMPLOYMENT_FULL else 'Part-time'

        for excel_file in request.FILES.getlist('excel_file'):
            if not excel_file.name.lower().endswith('.xlsx'):
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Only .xlsx files are allowed.",
                })
                continue

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "Could not open this file.",
                })
                continue

            if not rows:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "File is empty.",
                })
                continue

            header_error = check_header(rows[0], TEACHER_EXCEL_HEADERS)
            if header_error:
                file_results.append({
                    "filename": excel_file.name,
                    "error": header_error,
                })
                continue

            if len(rows) < 2:
                file_results.append({
                    "filename": excel_file.name,
                    "error": "This file appears to be empty (only a header row, no data rows).",
                })
                continue

            parsed_teachers = {}
            skipped = 0

            for row in rows[1:]:
                if not row or not any(cell is not None and str(cell).strip() != "" for cell in row):
                    continue
                if len(row) < 1:
                    skipped += 1
                    continue

                name = row[0]
                mobile = row[1] if len(row) > 1 else ""
                assigned_class = row[2] if len(row) > 2 else ""

                name_str = str(name).strip() if name is not None else ""
                clean_mobile = str(mobile).strip() if mobile else ""
                if clean_mobile.endswith('.0'):
                    clean_mobile = clean_mobile[:-2]
                if clean_mobile and not clean_mobile.startswith("0") and clean_mobile.isdigit():
                    clean_mobile = "0" + clean_mobile

                if not name_str or not clean_mobile:
                    skipped += 1
                    continue

                class_str = str(assigned_class).strip() if assigned_class else ""
                if class_str.endswith('.0'):
                    class_str = class_str[:-2]

                parsed_teachers[clean_mobile] = {
                    "name": name_str,
                    "mobile": clean_mobile,
                    "assigned_classes": class_str,
                }

            if not parsed_teachers:
                file_results.append({
                    "filename": excel_file.name,
                    "error": f"No valid rows found. All {skipped} row(s) were missing a Name or Number value.",
                })
                continue

            mobiles = list(parsed_teachers.keys())
            existing_users = {u.username: u for u in User.objects.filter(username__in=mobiles)}
            existing_teachers = {t.mobile: t for t in Teacher.objects.filter(mobile__in=mobiles).select_related('user')}

            created = 0
            updated = 0
            default_pw = getattr(settings, "DEFAULT_TEACHER_PASSWORD", "12345")

            with transaction.atomic():
                for mobile, data in parsed_teachers.items():
                    user = existing_users.get(mobile)
                    if not user:
                        user = User.objects.create_user(
                            username=mobile,
                            password=default_pw,
                        )
                        existing_users[mobile] = user

                    teacher = existing_teachers.get(mobile)
                    if teacher:
                        teacher.name = data["name"]
                        teacher.assigned_classes = data["assigned_classes"]
                        teacher.employment_type = employment_type
                        if teacher.user_id != user.id:
                            teacher.user = user
                        teacher.save()
                        updated += 1
                    else:
                        new_t = Teacher.objects.create(
                            name=data["name"],
                            mobile=mobile,
                            assigned_classes=data["assigned_classes"],
                            user=user,
                            employment_type=employment_type,
                        )
                        existing_teachers[mobile] = new_t
                        created += 1

            file_results.append({
                "filename": excel_file.name,
                "created": created,
                "updated": updated,
                "skipped": skipped,
            })

    return render(request, 'attendance/teacher_upload.html', {
        'file_results': file_results,
        'employment_type': employment_type,
        'type_label': type_label,
    })


@login_required
@user_passes_test(is_admin)
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    if request.method == 'POST':
        linked_user = teacher.user
        teacher.delete()
        if linked_user:
            linked_user.delete()
        return redirect('teacher_list')
    return render(request, 'attendance/teacher_confirm_delete.html', {'teacher': teacher})


@login_required
@user_passes_test(is_admin)
def mark_teacher_attendance(request):
    today = timezone.now().date()
    date_str = today.strftime("%d-%b-%y")
    saved = False
    saved_type = None
    sms_sent_count = 0
    sms_warning = None

    full_already_marked = TeacherAttendance.objects.filter(
        date=today, teacher__employment_type=Teacher.EMPLOYMENT_FULL
    ).exists()
    part_already_marked = TeacherAttendance.objects.filter(
        date=today, teacher__employment_type=Teacher.EMPLOYMENT_PART
    ).exists()

    if request.method == 'POST':
        section = request.POST.get('section')

        if section == 'full' and not full_already_marked:
            employment_type = Teacher.EMPLOYMENT_FULL
        elif section == 'part' and not part_already_marked:
            employment_type = Teacher.EMPLOYMENT_PART
        else:
            employment_type = None

        if employment_type:
            section_teachers = Teacher.objects.filter(employment_type=employment_type).order_by('id')
            absent_teachers = []
            for teacher in section_teachers:
                status = request.POST.get(f'tatt_{teacher.id}', 'present')
                is_present = status == 'present'

                TeacherAttendance.objects.update_or_create(
                    teacher=teacher,
                    date=today,
                    defaults={'is_present': is_present}
                )

                if not is_present:
                    absent_teachers.append(teacher)

            sms_sent_count, sms_failed = send_absent_sms(
                absent_teachers,
                build_teacher_absent_message,
                date_str,
                lambda teacher: teacher.mobile,
            )

            saved = True
            saved_type = section
            if section == 'full':
                full_already_marked = True
            else:
                part_already_marked = True
            if sms_failed:
                sms_warning = f"SMS could not be sent to: {', '.join(sms_failed)}"

    teachers = Teacher.objects.all().order_by('id')
    attendance_map = {
        a.teacher_id: a.is_present
        for a in TeacherAttendance.objects.filter(date=today)
    }
    teacher_rows = [
        {
            'teacher': t,
            'is_present': attendance_map.get(t.id, True),
        }
        for t in teachers
    ]

    full_time_rows = [r for r in teacher_rows if r['teacher'].employment_type == Teacher.EMPLOYMENT_FULL]
    part_time_rows = [r for r in teacher_rows if r['teacher'].employment_type == Teacher.EMPLOYMENT_PART]

    full_present = sum(1 for r in full_time_rows if r['is_present']) if full_already_marked else 0
    full_absent = len(full_time_rows) - full_present if full_already_marked else 0
    part_present = sum(1 for r in part_time_rows if r['is_present']) if part_already_marked else 0
    part_absent = len(part_time_rows) - part_present if part_already_marked else 0

    return render(request, 'attendance/mark_teacher_attendance.html', {
        'teachers': teachers,
        'teacher_rows': teacher_rows,
        'full_time_rows': full_time_rows,
        'part_time_rows': part_time_rows,
        'today': today,
        'saved': saved,
        'saved_type': saved_type,
        'sms_sent_count': sms_sent_count,
        'sms_warning': sms_warning,
        'full_already_marked': full_already_marked,
        'part_already_marked': part_already_marked,
        'full_total': len(full_time_rows),
        'part_total': len(part_time_rows),
        'full_present': full_present,
        'full_absent': full_absent,
        'part_present': part_present,
        'part_absent': part_absent,
        'total_teachers': teachers.count(),
    })

@login_required
@user_passes_test(is_admin)
def teacher_attendance_history(request):
    month_str = request.GET.get('month', '') or timezone.now().strftime('%Y-%m')
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        now = timezone.now()
        year, month = now.year, now.month
        month_str = f"{year:04d}-{month:02d}"

    teachers = Teacher.objects.all().order_by('id')
    records = []
    for teacher in teachers:
        qs = TeacherAttendance.objects.filter(teacher=teacher, date__year=year, date__month=month)
        present = qs.filter(is_present=True).count()
        absent = qs.filter(is_present=False).count()
        records.append({'teacher': teacher, 'present': present, 'absent': absent})

    return render(request, 'attendance/teacher_attendance_history.html', {
        'records': records,
        'month_str': month_str,
    })


@login_required
@user_passes_test(is_admin)
def export_teacher_attendance(request):
    month_str = request.GET.get('month', '') or timezone.now().strftime('%Y-%m')
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        now = timezone.now()
        year, month = now.year, now.month
        month_str = f"{year:04d}-{month:02d}"

    teachers = Teacher.objects.all().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Teacher Attendance {month_str}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3D32", end_color="1E3D32", fill_type="solid")
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["ID", "Name", "Mobile", "Present", "Absent"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_num, teacher in enumerate(teachers, 2):
        qs = TeacherAttendance.objects.filter(teacher=teacher, date__year=year, date__month=month)
        present = qs.filter(is_present=True).count()
        absent = qs.filter(is_present=False).count()

        values = [
            teacher.id,
            teacher.name,
            teacher.mobile or "",
            present,
            absent,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    filename = f"Teacher_Attendance_{month_str}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def attendance_history(request):
    class_names = Student.objects.values_list('class_name', flat=True).distinct().order_by('class_name')
    class_filter = request.GET.get('class', '').strip()
    roll_filter = request.GET.get('roll', '').strip()
    date_filter = request.GET.get('date', '') or timezone.now().date().isoformat()

    all_classes = [{'name': c, 'is_selected': (str(c) == class_filter)} for c in class_names]

    records = []
    if class_filter or roll_filter:
        students = Student.objects.all()
        if class_filter:
            students = students.filter(class_name=class_filter)
        if roll_filter:
            students = students.filter(roll_no__icontains=roll_filter)
        students = students.order_by('class_name', 'section', 'roll_no')

        attendance_map = {
            a.student_id: a.is_present
            for a in Attendance.objects.filter(student__in=students, date=date_filter)
        }
        for student in students:
            status = attendance_map.get(student.id)
            records.append({
                'student': student,
                'status': 'present' if status is True else ('absent' if status is False else 'not_marked'),
            })

    present_count = sum(1 for r in records if r['status'] == 'present')
    absent_count = sum(1 for r in records if r['status'] == 'absent')

    return render(request, 'attendance/attendance_history.html', {
        'all_classes': all_classes,
        'class_filter': class_filter,
        'roll_filter': roll_filter,
        'date_filter': date_filter,
        'records': records,
        'present_count': present_count,
        'absent_count': absent_count,
    })

@login_required
@user_passes_test(is_admin)
def export_attendance(request):
    class_filter = request.GET.get('class', '').strip()
    date_filter = request.GET.get('date', '') or timezone.now().date().isoformat()

    if not class_filter:
        return redirect('attendance_history')

    students = Student.objects.filter(class_name=class_filter).order_by('roll_no')
    attendance_map = {
        a.student_id: a.is_present
        for a in Attendance.objects.filter(student__class_name=class_filter, date=date_filter)
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Class {class_filter}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3D32", end_color="1E3D32", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Roll", "Name", "Class", "Section", "Mobile", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_num, student in enumerate(students, 2):
        status_bool = attendance_map.get(student.id)
        if status_bool is True:
            status = "Present"
            fill = present_fill
        elif status_bool is False:
            status = "Absent"
            fill = absent_fill
        else:
            status = "Not Marked"
            fill = None

        values = [
            student.roll_no,
            student.name,
            student.class_name,
            student.section,
            student.parent_mobile or "",
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
            if fill and col == 6:
                cell.fill = fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    filename = f"Attendance_Class_{class_filter}_{date_filter}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin)
def correct_attendance(request, student_id):
    """Admin-only: fix a wrongly marked attendance record for a given date."""
    if request.method != 'POST':
        return redirect('attendance_history')

    date_str = request.POST.get('date')
    class_filter = request.POST.get('class', '')
    roll_filter = request.POST.get('roll', '')
    new_status = request.POST.get('status')  # 'present' or 'absent'
    send_sms_flag = request.POST.get('send_sms') == 'yes'

    redirect_url = f"{reverse('attendance_history')}?class={class_filter}&roll={roll_filter}&date={date_str}"

    if new_status not in ('present', 'absent'):
        return redirect(redirect_url)

    student = get_object_or_404(Student, id=student_id)
    Attendance.objects.update_or_create(
        student=student,
        date=date_str,
        defaults={'is_present': new_status == 'present'}
    )

    if new_status == 'absent' and send_sms_flag and student.parent_mobile:
        message = build_absent_message(student, date_str)
        send_sms(student.parent_mobile, message)

    return redirect(redirect_url)